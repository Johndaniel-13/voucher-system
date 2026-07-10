"""
EXCEL PARSER — Reads admin-uploaded Excel files
================================================
Expected Excel format:
  Column A: Serial No.
  Column B: Voucher Code

Also handles exporting issued vouchers as Excel.
"""

import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def parse_voucher_excel(file_bytes: bytes) -> list[dict]:
    """
    Parse admin's voucher Excel file.

    Expected columns:
      - Serial No. (integer)
      - Voucher Code (string)

    Returns list of {"serial_no": int, "voucher_code": str}
    """
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes))
    sheet = workbook.active

    vouchers = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header row
        if row[0] and row[1]:  # Both columns must have data
            vouchers.append({
                "serial_no": int(row[0]),
                "voucher_code": str(row[1]).strip()
            })

    workbook.close()
    return vouchers


def export_issued_to_excel(issued_data: list[dict]) -> bytes:
    """
    Create an Excel file from issued vouchers data.
    Columns: Payment UTR Number | Voucher Code | Issued At

    Returns Excel file as bytes.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Issued Vouchers"

    # Headers with styling
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    headers = ["S.No", "Payment UTR Number", "Voucher Code", "Issued At"]

    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Data rows
    for idx, record in enumerate(issued_data, 1):
        sheet.cell(row=idx + 1, column=1, value=idx)
        sheet.cell(row=idx + 1, column=2, value=record.get("utr_number", ""))
        sheet.cell(row=idx + 1, column=3, value=record.get("voucher_code", ""))
        sheet.cell(row=idx + 1, column=4, value=str(record.get("issued_at", "")))

    # Auto-adjust column widths
    for col in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        sheet.column_dimensions[col[0].column_letter].width = max_length + 5

    # Save to bytes
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    workbook.close()

    return output.getvalue()
