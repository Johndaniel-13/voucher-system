"""
Create a sample Excel file with test voucher codes.
Run this to generate: python create_sample_excel.py
"""

import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Voucher Codes"

# Header
ws["A1"] = "Serial No."
ws["B1"] = "Voucher Code"
ws["A1"].font = openpyxl.styles.Font(bold=True)
ws["B1"].font = openpyxl.styles.Font(bold=True)

# Sample data
for i in range(1, 21):
    ws.append([i, f"VOUCH-{i:03d}-RAJA"])

wb.save("sample_vouchers.xlsx")
print("✅ Created sample_vouchers.xlsx with 20 voucher codes!")
